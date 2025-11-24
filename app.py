#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2025/11/5 16:23
# @Author  : Healer
# @File    : app.py
# @Software: PyCharm

import streamlit as st
import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string
from datetime import datetime
import tempfile
import shutil
from copy import copy
import zipfile
from io import BytesIO
import hashlib
import re


class SIGeneratorWeb:
    def __init__(self):
        self.setup_page()

    def setup_page(self):
        st.set_page_config(
            page_title="SI Generator Tool",
            page_icon="📊",
            layout="wide",
            initial_sidebar_state="expanded"
        )

        st.title("📊 SI Generator Tool Made by Kexue")
        st.markdown("---")

    def run(self):
        # Sidebar for file upload
        with st.sidebar:
            st.header("Upload Files")
            uploaded_files = st.file_uploader(
                "Choose Excel files",
                type=['xlsx', 'xlsm'],
                accept_multiple_files=True,
                help="Upload one or more Excel files containing 'No SI Order' and 'SI Template' sheets"
            )

            st.header("Settings")
            group_column = st.selectbox(
                "Group by Column",
                ["O", "P", "Q", "R", "S", "T"],
                index=0,
                help="Select the column to group data by"
            )

            # 添加输出目录选择
            output_option = st.radio(
                "Output Location",
                ["Same as input files", "Download only"],
                help="Choose where to save the generated files"
            )

            if st.button("Generate SI Files", type="primary", use_container_width=True):
                if uploaded_files:
                    self.process_files(uploaded_files, group_column, output_option)
                else:
                    st.error("Please upload at least one Excel file")

        # Main area for instructions
        st.markdown("""
        ### 📋 Instructions

        1. **Upload Excel Files**: Select one or more Excel files (.xlsx or .xlsm) that contain:
           - 'No SI Order' sheet with data
           - 'SI Template' sheet with formatting

        2. **Select Group Column**: Choose which column to group the data by (default is Column O)

        3. **Choose Output Location**: 
           - **Same as input files**: Files will be saved in the same folder as your original Excel files
           - **Download only**: Files will only be available for download

        4. **Generate**: Click the 'Generate SI Files' button to process all files

        ### 📁 Output Structure
        For each input file, the tool will create:
        - Individual SI files for each group
        - A consolidated file with all SIs
        - All files will be saved in a folder next to your original file
        """)

    def process_files(self, uploaded_files, group_column, output_option):
        progress_bar = st.progress(0)
        status_text = st.empty()
        log_container = st.container()

        total_files = len(uploaded_files)
        processed_files = 0

        # 如果选择"Download only"，我们仍然需要临时目录
        if output_option == "Download only":
            # Create a ZIP file to store all results
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:

                for i, uploaded_file in enumerate(uploaded_files):
                    progress = i / total_files
                    progress_bar.progress(progress)
                    status_text.text(f"Processing {i + 1}/{total_files}: {uploaded_file.name}")

                    try:
                        # Process single file - 下载模式下使用临时目录
                        result_files = self.process_single_file_download(uploaded_file, group_column, log_container)

                        # Add files to ZIP
                        for file_path in result_files:
                            zip_file.write(file_path, os.path.basename(file_path))

                        processed_files += 1
                        log_container.success(f"✅ Successfully processed: {uploaded_file.name}")

                    except Exception as e:
                        log_container.error(f"❌ Error processing {uploaded_file.name}: {str(e)}")
                        continue

                progress_bar.progress(1.0)
                status_text.text("Processing completed!")

            # Provide download link
            if processed_files > 0:
                zip_buffer.seek(0)
                st.download_button(
                    label="📥 Download All SI Files",
                    data=zip_buffer,
                    file_name=f"SI_Files_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                    mime="application/zip",
                    use_container_width=True
                )

                st.success(f"🎉 Successfully processed {processed_files}/{total_files} files!")
            else:
                st.error("No files were successfully processed.")

        else:  # Same as input files
            for i, uploaded_file in enumerate(uploaded_files):
                progress = i / total_files
                progress_bar.progress(progress)
                status_text.text(f"Processing {i + 1}/{total_files}: {uploaded_file.name}")

                try:
                    # Process single file - 保存到原始文件目录
                    result_files = self.process_single_file_local(uploaded_file, group_column, log_container)

                    processed_files += 1
                    log_container.success(f"✅ Successfully processed: {uploaded_file.name}")

                    # 显示保存路径
                    if result_files:
                        first_file_dir = os.path.dirname(result_files[0])
                        log_container.info(f"   📁 Files saved to: {first_file_dir}")

                except Exception as e:
                    log_container.error(f"❌ Error processing {uploaded_file.name}: {str(e)}")
                    continue

            progress_bar.progress(1.0)
            status_text.text("Processing completed!")

            if processed_files > 0:
                st.success(f"🎉 Successfully processed {processed_files}/{total_files} files!")
                st.info("💡 Generated files have been saved in the same folders as your original Excel files.")
            else:
                st.error("No files were successfully processed.")

    def process_single_file_download(self, uploaded_file, group_column, log_container):
        """处理文件并返回文件路径（下载模式）"""
        # 创建临时目录
        with tempfile.TemporaryDirectory() as temp_dir:
            return self._process_single_file(uploaded_file, group_column, log_container, temp_dir)

    def process_single_file_local(self, uploaded_file, group_column, log_container):
        """处理文件并保存到原始文件目录"""
        # 获取用户桌面路径作为默认保存位置
        try:
            save_dir = self.get_save_directory()
            return self._process_single_file(uploaded_file, group_column, log_container, save_dir)
        except Exception as e:
            log_container.warning(f"⚠️ Cannot access original file location, using download mode instead: {str(e)}")
            return self.process_single_file_download(uploaded_file, group_column, log_container)

    def get_save_directory(self):
        """获取保存目录"""
        try:
            # 尝试获取桌面路径
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            if os.path.exists(desktop):
                return desktop
            else:
                return os.getcwd()
        except:
            return os.getcwd()

    def _process_single_file(self, uploaded_file, group_column, log_container, base_dir):
        """处理单个文件的通用逻辑"""
        # 创建临时文件来处理上传的内容
        temp_file_path = os.path.join(base_dir, f"temp_{uploaded_file.name}")
        with open(temp_file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        try:
            # Load workbook
            workbook = load_workbook(temp_file_path, data_only=True)

            # Check required sheets
            if 'No SI Order' not in workbook.sheetnames:
                raise Exception("'No SI Order' sheet not found")

            if 'SI Template' not in workbook.sheetnames:
                raise Exception("'SI Template' sheet not found")

            # Get data from sheets
            order_sheet = workbook['No SI Order']
            template_sheet = workbook['SI Template']

            # Group data
            grouped_data = self.group_data_by_column(order_sheet, group_column)

            if not grouped_data:
                raise Exception(f"No valid data found in column {group_column}")

            log_container.info(f"📁 {uploaded_file.name}: Found {len(grouped_data)} groups")

            # 创建输出目录 - 使用完整的原始文件名
            file_base_name = os.path.splitext(uploaded_file.name)[0]
            output_folder = os.path.join(base_dir, f"SI_Output_{file_base_name}")
            os.makedirs(output_folder, exist_ok=True)

            result_files = []
            si_count = 0

            # Create consolidated workbook
            consolidated_wb = openpyxl.Workbook()
            consolidated_wb.remove(consolidated_wb.active)

            for group_key, rows in grouped_data.items():
                try:
                    # Create individual SI file
                    individual_si_path = self.create_individual_si(
                        group_key, rows, order_sheet, template_sheet, output_folder, uploaded_file.name
                    )
                    result_files.append(individual_si_path)

                    # Add to consolidated workbook
                    self.add_to_consolidated_workbook(
                        group_key, rows, order_sheet, template_sheet, consolidated_wb
                    )

                    si_count += 1
                    log_container.info(f"   ✅ Created SI for group: {group_key}")

                except Exception as e:
                    log_container.warning(f"   ⚠️ Skipped group {group_key}: {str(e)}")
                    continue

            # Save consolidated workbook - 使用完整文件名
            if si_count > 0:
                consolidated_filename = f"Consolidated_SI_{file_base_name}.xlsx"
                consolidated_path = os.path.join(output_folder, consolidated_filename)
                consolidated_wb.save(consolidated_path)
                result_files.append(consolidated_path)

            return result_files

        finally:
            # 清理临时文件
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)

    def group_data_by_column(self, sheet, column_letter):
        """Group data by specified column"""
        grouped_data = {}

        try:
            col_idx = column_index_from_string(column_letter)

            for row in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row, column=col_idx).value
                if cell_value and str(cell_value).strip():
                    key = str(cell_value).strip()
                    if key not in grouped_data:
                        grouped_data[key] = []
                    grouped_data[key].append(row)

        except Exception as e:
            raise Exception(f"Error grouping data: {str(e)}")

        return grouped_data

    def copy_sheet_with_formatting(self, source_sheet, target_sheet):
        """Copy entire sheet with all formatting"""
        try:
            # Copy column dimensions
            for col_letter, col_dim in source_sheet.column_dimensions.items():
                target_sheet.column_dimensions[col_letter].width = col_dim.width
                if col_dim.hidden:
                    target_sheet.column_dimensions[col_letter].hidden = True

            # Copy row dimensions
            for row_num, row_dim in source_sheet.row_dimensions.items():
                target_sheet.row_dimensions[row_num].height = row_dim.height
                if row_dim.hidden:
                    target_sheet.row_dimensions[row_num].hidden = True

            # Copy merged cells
            for merged_range in source_sheet.merged_cells.ranges:
                target_sheet.merge_cells(str(merged_range))

            # Copy all cells with formatting
            for row in source_sheet.iter_rows():
                for cell in row:
                    new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

                    if cell.hyperlink:
                        new_cell.hyperlink = cell.hyperlink
                        new_cell.style = "Hyperlink"

            target_sheet.page_setup = copy(source_sheet.page_setup)

            if hasattr(source_sheet, 'print_options'):
                target_sheet.print_options = copy(source_sheet.print_options)

        except Exception as e:
            raise Exception(f"Error copying sheet formatting: {str(e)}")

    def create_individual_si(self, group_key, rows, order_sheet, template_sheet, output_folder, original_filename):
        """Create individual SI Excel file for a group"""
        si_wb = openpyxl.Workbook()
        si_wb.remove(si_wb.active)

        si_sheet = si_wb.create_sheet(title="SI")
        self.copy_sheet_with_formatting(template_sheet, si_sheet)

        # Fill specific information
        self.fill_specific_info(order_sheet, si_sheet, rows[0])

        # Fill table data
        self.fill_table_data(order_sheet, si_sheet, rows)

        # 使用完整的原始文件名，不再截断
        safe_group_key = self.create_safe_filename(group_key)
        file_base_name = os.path.splitext(original_filename)[0]  # 不再限制长度

        # 使用完整的文件名格式
        si_filename = f"SI_{safe_group_key}_{file_base_name}.xlsx"
        si_path = os.path.join(output_folder, si_filename)

        si_wb.save(si_path)

        return si_path

    def create_safe_filename(self, text):
        """创建安全的文件名，移除特殊字符但保持完整内容"""
        # 只移除Windows不允许的特殊字符，保持其他内容完整
        safe_text = re.sub(r'[<>:"/\\|?*]', '', text)
        safe_text = re.sub(r'\s+', ' ', safe_text.strip())  # 保留空格，不替换为下划线
        return safe_text  # 不再限制长度

    def fill_specific_info(self, order_sheet, si_sheet, first_data_row):
        """Fill specific information into SI template"""
        field_mapping = [
            ("Destination Port", "B1"),
            ("Customer Name", "B2"),
            ("Payment Term", "B3"),
            ("Transport Mode", "B5"),
            ("Incoterm", "B6"),
            ("Forwarder Name", "B13"),
            ("Forwarder Contact Person", "B14"),
            ("Forwarder Telephone No", "B15"),
            ("Forwarder E-Mail", "B16")
        ]

        for field_name, cell_address in field_mapping:
            col_idx = self.find_column_index(order_sheet, field_name)
            if col_idx:
                value = order_sheet.cell(row=first_data_row, column=col_idx).value
                si_sheet[cell_address].value = value

    def fill_table_data(self, order_sheet, si_sheet, data_rows):
        """Fill table data starting from row 19"""
        columns_to_copy = [
            "Order Nbr", "Total Qty", "Shipment Wt", "Volumetric Wt",
            "Age (Days)", "PO Nbr", "Sales Rep Name"
        ]

        start_row = 19

        for i, source_row in enumerate(data_rows):
            target_row = start_row + i

            if target_row > si_sheet.max_row:
                for col in range(1, len(columns_to_copy) + 1):
                    source_cell = si_sheet.cell(row=target_row - 1, column=col)
                    target_cell = si_sheet.cell(row=target_row, column=col)

                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)

            for col_idx, column_name in enumerate(columns_to_copy, 1):
                source_col_idx = self.find_column_index(order_sheet, column_name)
                if source_col_idx:
                    value = order_sheet.cell(row=source_row, column=source_col_idx).value
                    si_sheet.cell(row=target_row, column=col_idx).value = value

    def add_to_consolidated_workbook(self, group_key, rows, order_sheet, template_sheet, consolidated_wb):
        """Add SI data to consolidated workbook"""
        # 使用完整的组名作为工作表名称
        safe_sheet_name = "".join(c for c in str(group_key) if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_sheet_name = f"SI_{safe_sheet_name}"  # 不再截断

        si_sheet = consolidated_wb.create_sheet(title=safe_sheet_name)
        self.copy_sheet_with_formatting(template_sheet, si_sheet)

        self.fill_specific_info(order_sheet, si_sheet, rows[0])
        self.fill_table_data(order_sheet, si_sheet, rows)

    def find_column_index(self, sheet, column_name):
        """Find column index by header name"""
        for cell in sheet[1]:
            if cell.value and str(cell.value).strip().lower() == column_name.lower():
                return cell.column
        return None


def main():
    app = SIGeneratorWeb()
    app.run()


if __name__ == "__main__":
    main()